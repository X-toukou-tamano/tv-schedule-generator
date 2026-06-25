import os
import re
import unicodedata
from calendar import monthrange
from datetime import date
from openpyxl import load_workbook

KEIRIN_TRACKS = [
    "函館", "青森", "いわき平", "弥彦", "前橋", "取手", "宇都宮", "大宮", "西武園", "京王閣", "立川",
    "松戸", "川崎", "平塚", "小田原", "伊東", "静岡", "名古屋", "岐阜", "大垣", "豊橋", "富山",
    "松阪", "四日市", "福井", "奈良", "向日町", "和歌山", "岸和田", "玉野", "広島", "防府", "高松", "小松島",
    "高知", "松山", "小倉", "久留米", "武雄", "佐世保", "別府", "熊本"
]

TARGET_BLOCKS = [
    "玉野",
    "現金機＆CLAP"
]

IGNORE_VALUES = [
    "開催なし",
    "発売中止",
    "その他",
    "備考",
    "非開催"
]

def clean_block_name(name):
    if not isinstance(name, str):
        return name
    name = unicodedata.normalize('NFKC', name)
    name = name.replace(" ", "").replace(" ", "").strip()
    name = name.replace("&", "＆")
    return name

def build_merged_map(ws):
    merged_map = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        parent_val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = parent_val
    return merged_map

def get_merged_value(cell, merged_map):
    if cell.value is not None:
        return cell.value
    return merged_map.get((cell.row, cell.column), None)

def find_months(ws, merged_map):
    months = {}
    for row in ws.iter_rows():
        for cell in row:
            value = get_merged_value(cell, merged_map)
            if not isinstance(value, str):
                continue
            
            value = unicodedata.normalize('NFKC', value).replace(" ", "").replace(" ", "").strip()
            
            match = re.match(r'^([1-9]|1[0-2])月$', value)
            if match:
                month_num = int(match.group(1))
                if month_num not in months:
                    months[month_num] = cell
    return months

def get_day1_column(ws, month_cell):
    for merged in ws.merged_cells.ranges:
        if month_cell.coordinate in merged:
            return merged.max_col + 1
    return month_cell.column + 1

def resolve_year(filename, month):
    fiscal_year = 2026 
    match_year = re.search(r'(20\d{2})', filename)
    if match_year:
        fiscal_year = int(match_year.group(1))
    else:
        match_r = re.search(r'R([0-9０-９]+)', filename, re.IGNORECASE)
        if match_r:
            r_num = int(match_r.group(1).translate(str.maketrans('０１２３４５６７８９', '0123456789')))
            fiscal_year = 2018 + r_num

    if month >= 4:
        return fiscal_year
    return fiscal_year + 1

def find_block_column(ws, merged_map):
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, 10):
            val = get_merged_value(ws.cell(row, col), merged_map)
            cleaned_val = clean_block_name(val)
            if isinstance(cleaned_val, str) and cleaned_val in TARGET_BLOCKS:
                return col
    return 2 

def normalize_venue_name(raw_name):
    if not isinstance(raw_name, str):
        return None
        
    name = unicodedata.normalize('NFKC', raw_name)
    name_no_space = name.replace(" ", "").replace(" ", "")

    if re.match(r'^\d+-\d+', name_no_space):
        return "玉野"

    for track in KEIRIN_TRACKS:
        if track in name_no_space:
            return track

    return None

def extract_venues(ws, target_col, block_col, merged_map, start_row, end_row):
    venues = []
    current_block = None

    for row in range(start_row, end_row + 1):
        block_name = get_merged_value(ws.cell(row, block_col), merged_map)
        cleaned_block = clean_block_name(block_name)

        if isinstance(cleaned_block, str) and cleaned_block != "":
            if cleaned_block in TARGET_BLOCKS:
                current_block = cleaned_block
            else:
                current_block = None
        else:
            if current_block is not None and current_block not in TARGET_BLOCKS:
                current_block = None

        if current_block not in TARGET_BLOCKS:
            continue

        value = get_merged_value(ws.cell(row, target_col), merged_map)

        if value is None:
            continue

        value_str = str(value).strip()
        value_no_space = value_str.replace(" ", "").replace(" ", "")

        if value_str == "" or value_str in TARGET_BLOCKS:
            continue

        if any(ignore_word in value_no_space for ignore_word in IGNORE_VALUES):
            continue

        cleaned_venue = normalize_venue_name(value_str)
        if cleaned_venue:
            venues.append(cleaned_venue)

    return venues

def parse_excel(excel_path):
    filename = os.path.basename(excel_path)
    print(filename)
    wb = load_workbook(excel_path, data_only=True)
    records = []
    
    for ws in wb.worksheets:
        # 隠しシートは絶対に読み込まない（前回のバグの確実な防止）
        if ws.sheet_state != 'visible':
            continue
            
        merged_map = build_merged_map(ws)
        month_map = find_months(ws, merged_map)
        
        if not month_map:
            continue
            
        block_col = find_block_column(ws, merged_map)
        sorted_months = sorted(month_map.items(), key=lambda item: (item[1].row, item[1].column))

        for i, (month, month_cell) in enumerate(sorted_months):
            year = resolve_year(filename, month)
            day1_col = get_day1_column(ws, month_cell)
            days_in_month = monthrange(year, month)[1]

            start_row = month_cell.row
            
            end_row = ws.max_row
            for j in range(i + 1, len(sorted_months)):
                next_month_cell = sorted_months[j][1]
                if next_month_cell.row > start_row:
                    end_row = next_month_cell.row - 1
                    break

            for day in range(1, days_in_month + 1):
                target_col = day1_col + day - 1
                venues = extract_venues(ws, target_col, block_col, merged_map, start_row, end_row)

                for venue in venues:
                    records.append({
                        "date": date(year, month, day),
                        "venue": venue
                    })
                    
    return records

def get_upload_info(excel_path):

    filename = os.path.basename(excel_path)

    # ----------------------------
    # 上期・下期判定
    # ----------------------------

    wb = load_workbook(
        excel_path,
        data_only=True
    )

    term = None

    for ws in wb.worksheets:
        print(type(ws))
        print(ws.__class__)
        print(hasattr(ws, "merged_cells"))
        if ws.sheet_state != "visible":
            continue

        merged_map = build_merged_map(ws)
        month_map = find_months(
            ws,
            merged_map
        )

        if not month_map:
            continue

        first_month = min(
            month_map.keys()
        )

        if 4 <= first_month <= 9:
            term = "上期"
        else:
            term = "下期"

        break

    wb.close()

    if term is None:
        raise ValueError(
            "上期・下期を判定できません。"
        )

    # ----------------------------
    # 年度判定
    # ----------------------------

    match_r = re.search(
        r"R([0-9０-９]+)",
        filename,
        re.IGNORECASE
    )

    if match_r:

        year = (
            "R"
            + str(
                int(
                    match_r.group(1).translate(
                        str.maketrans(
                            "０１２３４５６７８９",
                            "0123456789"
                        )
                    )
                )
            )
        )

    else:

        match_year = re.search(
            r"(20\d{2})",
            filename
        )

        if not match_year:
            raise ValueError(
                "年度を判定できません。"
            )

        fiscal_year = int(
            match_year.group(1)
        )

        # 西暦は年度で判定
        if term == "上期":
            reiwa = fiscal_year - 2018
        else:
            reiwa = fiscal_year - 2019

        year = f"R{reiwa}"

    return (
        year,
        term
    )
